"""
Spreadsheet skill: create or update XLSX and CSV files.
Manus-style "delivers spreadsheets".
"""
import os
import csv
from typing import Dict, Any, List

from viki.skills.base import BaseSkill
from viki.config.logger import viki_logger
from viki.core.utils.path_sandbox import validate_output_path


class SpreadsheetSkill(BaseSkill):
    def __init__(self, controller=None):
        self._controller = controller

    @property
    def name(self) -> str:
        return "spreadsheet"

    @property
    def description(self) -> str:
        return (
            "Create or update spreadsheets (XLSX or CSV). Actions: create_spreadsheet(path=..., headers=[...], rows=[...]), "
            "create_from_data(path=..., data=[{...}, ...]), append_rows(path=..., rows=[...])."
        )

    def _ensure_dir(self, path: str) -> None:
        d = os.path.dirname(path)
        if d:
            os.makedirs(d, exist_ok=True)

    async def execute(self, params: Dict[str, Any]) -> str:
        path = params.get("path") or params.get("output_path") or params.get("file")
        if not path:
            return "Provide path= (output file path, .xlsx or .csv)."
        ok, path_or_err = validate_output_path(path, controller=self._controller)
        if not ok:
            return path_or_err

        action = params.get("action", "create_from_data")
        path = path_or_err
        self._ensure_dir(path)
        is_csv = path.lower().endswith(".csv")

        if action == "append_rows":
            rows = params.get("rows") or params.get("data")
            if not rows:
                return "append_rows requires rows= (list of lists or list of dicts)."
            try:
                if is_csv:
                    with open(path, "a", newline="", encoding="utf-8") as f:
                        w = csv.writer(f)
                        for row in rows:
                            w.writerow(row if isinstance(row, (list, tuple)) else list(row.values()))
                else:
                    from openpyxl import load_workbook
                    wb = load_workbook(path)
                    ws = wb.active
                    for row in rows:
                        r = row if isinstance(row, (list, tuple)) else list(row.values())
                        ws.append(r)
                    wb.save(path)
                return f"Appended {len(rows)} row(s) to {path}."
            except FileNotFoundError:
                return f"File not found: {path}. Create it first with create_spreadsheet or create_from_data."
            except Exception as e:
                return f"append_rows error: {e}"

        # create_spreadsheet: headers + rows
        headers = params.get("headers")
        rows = params.get("rows")
        if headers is not None:
            if rows is None:
                rows = []
            try:
                if is_csv:
                    with open(path, "w", newline="", encoding="utf-8") as f:
                        w = csv.writer(f)
                        w.writerow(headers)
                        w.writerows(rows)
                else:
                    from openpyxl import Workbook
                    wb = Workbook()
                    ws = wb.active
                    ws.append(headers)
                    for row in rows:
                        if isinstance(row, (list, tuple)):
                            ws.append(list(row))
                        else:
                            ws.append([row.get(h) for h in headers])
                    wb.save(path)
                return f"Spreadsheet saved to {path}."
            except Exception as e:
                return f"Error: {e}"

        # create_from_data: list of dicts
        data = params.get("data")
        if not data or not isinstance(data, list):
            return "Provide headers= and rows=, or data= (list of dicts)."
        try:
            if is_csv:
                if not data:
                    return "data= is empty."
                keys = list(data[0].keys())
                with open(path, "w", newline="", encoding="utf-8") as f:
                    w = csv.DictWriter(f, fieldnames=keys)
                    w.writeheader()
                    w.writerows(data)
            else:
                from openpyxl import Workbook
                wb = Workbook()
                ws = wb.active
                keys = list(data[0].keys())
                ws.append(keys)
                for row in data:
                    ws.append([row.get(k) for k in keys])
                wb.save(path)
            return f"Spreadsheet saved to {path}."
        except Exception as e:
            viki_logger.warning(f"spreadsheet create: {e}")
            return f"Error: {e}"
